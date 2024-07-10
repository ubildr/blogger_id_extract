import streamlit as st
import pandas as pd
import requests
import time
from datetime import datetime
from urllib.parse import urlparse
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def get_blogger_ids(keyword, text_area, max_results=1000):
    client_id = st.secrets["NAVER_CLIENT_ID"]
    client_secret = st.secrets["NAVER_CLIENT_SECRET"]
    url = "https://openapi.naver.com/v1/search/blog.json"
    headers = {
        "X-Naver-Client-Id": client_id,
        "X-Naver-Client-Secret": client_secret
    }

    blogger_ids = set()
    start = 1
    total_pages = 10

    for page in range(1, total_pages + 1):
        if len(blogger_ids) >= max_results:
            break

        params = {
            "query": keyword,
            "display": 100,
            "start": start
        }

        try:
            response = requests.get(url, headers=headers, params=params)
            response.raise_for_status()

            data = response.json()
            items = data.get('items', [])

            if not items:
                break

            for item in items:
                parsed_url = urlparse(item['bloggerlink'])
                blogger_id = parsed_url.path.strip('/').split('/')[-1]
                if blogger_id:
                    blogger_ids.add(blogger_id)

            start += 100
            time.sleep(0.1)

            text_area.text(f"({page}/{total_pages}) {len(blogger_ids)}개의 블로거 ID 추출 완료")

        except requests.exceptions.RequestException as e:
            text_area.error(f"API 요청 중 오류 발생, 추출 중단")
            break

    return list(blogger_ids)[:max_results]


st.title("네이버 블로거 ID 추출 앱")

keywords = st.text_input("검색어를 입력하세요 (콤마 또는 세미콜론으로 구분):")

if st.button("블로거 ID 추출"):
    if keywords:
        keywords_list = [k.strip() for k in keywords.replace(';', ',').split(',')]
        all_data = []

        progress_bar = st.progress(0)
        text_area = st.empty()

        for i, keyword in enumerate(keywords_list, 1):
            text_area.text(f"[{i}/{len(keywords_list)}] '{keyword}' 검색어 처리 중...")
            with st.spinner(f"[{i}/{len(keywords_list)}] '{keyword}'에 대한 블로거 ID를 추출 중입니다..."):
                blogger_ids = get_blogger_ids(keyword, text_area=text_area)
            
            df = pd.DataFrame({"검색어": [keyword] * len(blogger_ids), "블로거 ID": blogger_ids})
            all_data.append(df)
            
            progress_bar.progress(i / len(keywords_list))

        final_df = pd.concat(all_data, ignore_index=True)
        st.success(f"총 {len(final_df)}개의 고유한 블로거 ID를 추출했습니다.")
        
        st.dataframe(final_df)
        
        st.session_state.df = final_df
    else:
        st.warning("검색어를 입력해주세요.")

if 'df' in st.session_state:
    current_date = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    excel_filename = f"blogger_ids_{current_date}.xlsx"
    
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    
    # 헤더 추가 및 스타일 설정
    headers = ["키워드", "블로거 ID"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # 데이터 추가
    for r, row in enumerate(st.session_state.df.values, 2):
        for c, value in enumerate(row, 1):
            ws.cell(row=r, column=c, value=value)
    
    # 필터 적용
    ws.auto_filter.ref = ws.dimensions
    
    # A2 셀 고정
    ws.freeze_panes = 'A2'
    
    wb.save(output)
    output.seek(0)
    
    st.download_button(
        label="엑셀 파일 다운로드",
        data=output.getvalue(),
        file_name=excel_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
else:
    st.warning("먼저 블로거 ID를 추출해주세요.")